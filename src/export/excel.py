"""Excel workbook export with formatting."""

from typing import List, Dict, Optional
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelExporter:
    """Exports session data to formatted Excel workbook."""

    HEADER_FILL = PatternFill(start_color='174A7E', end_color='174A7E', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF')

    def __init__(self, assessor_name: str):
        """Initialize Excel exporter."""
        self.assessor_name = assessor_name
        self.summary_data: List[Dict] = []
        self.flags_data: List[Dict] = []
        self.time_logs_data: List[Dict] = []

    def add_session_summary(
        self,
        committed_count: int,
        flags_count: int,
        avg_minutes: float,
        baseline_minutes: Optional[float] = None,
        total_minutes: float = 0.0,
    ) -> None:
        """Add session summary data."""
        self.summary_data = [{
            'Assessor': self.assessor_name,
            'Claims Committed': committed_count,
            'Flags Count': flags_count,
            'Average Minutes': round(avg_minutes, 2),
            'Baseline Minutes': baseline_minutes,
            'Total Minutes': round(total_minutes, 2),
            'Exported At': datetime.now().isoformat(timespec='seconds'),
        }]

    def add_flagged_claims(self, flags: List[Dict]) -> None:
        """Add flagged claims data."""
        self.flags_data = flags

    def add_time_logs(self, time_logs: List[Dict]) -> None:
        """Add time log data."""
        self.time_logs_data = time_logs

    def export(self, output_path: str) -> Path:
        """Export to Excel workbook."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        summary_df = pd.DataFrame(self.summary_data)
        flags_df = pd.DataFrame(self.flags_data) if self.flags_data else pd.DataFrame()
        time_logs_df = pd.DataFrame(self.time_logs_data) if self.time_logs_data else pd.DataFrame()

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, index=False, sheet_name='Session Summary')
            if not flags_df.empty:
                flags_df.to_excel(writer, index=False, sheet_name='Flagged Claims')
            if not time_logs_df.empty:
                time_logs_df.to_excel(writer, index=False, sheet_name='Time Logs')

        self._format_workbook(output_path)
        return output_path

    def _format_workbook(self, file_path: Path) -> None:
        """Apply formatting to workbook."""
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            ws.freeze_panes = 'A2'
            for cell in ws[1]:
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(45, max(12, max_length + 2))
                ws.column_dimensions[column_letter].width = adjusted_width
            if ws.max_row > 0 and ws.max_column > 0:
                ws.auto_filter.ref = ws.dimensions
        wb.save(file_path)
